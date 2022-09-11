import json
from pprint import pprint
from re import sub
from string import ascii_uppercase
import string
from unittest import FunctionTestCase
from selenium import webdriver
import selenium.common.exceptions as exc
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium_stealth import stealth
import time
from typing import (
    Iterable,
    Mapping,
)
import pandas as pd
from ParserScripts.areas import serialize_data_areas
from ParserScripts.categories import get_categories
from selenium.webdriver.support.ui import Select
import os
from bs4 import BeautifulSoup
from datetime import date, datetime
from openpyxl import Workbook
from functions.logger import get_logger
from functions.logger import file_name_datetime
LOGGER = get_logger()
class Avito_Posts_Parser:
    """
    ____ _  _ _ ___ ____     ___  ____ ____ ___ ____     ___  ____ ____ ____ ____ ____ 
    |__| |  | |  |  |  |     |__] |  | [__   |  [__      |__] |__| |__/ [__  |___ |__/ 
    |  |  \/  |  |  |__| ___ |    |__| ___]  |  ___] ___ |    |  | |  \ ___] |___ |  \ 
    """
    def __init__(self, city: str, parameters: Mapping) -> None:
        """
        Инициализируем класс парсера
        """
        self.city: str = city
        self.parameters: Iterable = parameters
        self.__LINK: str = "https://www.avito.ru"
        self.__DEBUG = os.environ.get("DEBUG")
        self.__driver: webdriver.Chrome() = None

    def __configure_selenium(self) -> None:
        """
        Настраивает селениум для правильной работы
        """
        LOGGER.info(f"Функция __configure_selenium запущена {datetime.now()}")
        try:        
            options = webdriver.ChromeOptions()
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option('useAutomationExtension', False)
            options.headless = True
            options.add_argument("--headless")
            options.add_argument("--disable-extensions")
            executable_path = os.environ['EXECUTABLE_PATH']
            driver = webdriver.Chrome(options=options, executable_path=executable_path)
            stealth(
                    driver,
                    languages=["en-US", "en"],
                    vendor="Google Inc.",
                    platform="Win32",
                    webgl_vendor="Intel Inc.",
                    renderer="Intel Iris OpenGL Engine",
                    fix_hairline=True,
            )
            self.__driver: driver = driver
            self.__driver.delete_all_cookies()

        except Exception as e:
            print(e, flush=True) 
            LOGGER.critical(f"Функция __configure_selenium была остановлена по причине ошибки {e} {datetime.now()}")


    def run(self):
        """
        Запускает скрипт со всеми сценариями
        """
        LOGGER.info(f"Функция run запущена {datetime.now()}") 
        self.__configure_selenium()
        self.__driver.get(self.__LINK)   
        time.sleep(self.parameters["wait_time"])
        self.__make_clicks_to_get_right_city()
        time.sleep(self.parameters["wait_time"])
        self.__only_name_choose()
        time.sleep(self.parameters["wait_time"])
        self.__make_clicks_to_get_category()
        time.sleep(self.parameters["wait_time"])
        first_page_link = self.__driver.current_url
        final_dict_product = self.parse_content_on_page(self.__driver.page_source)
        self.dict_to_json(dict_=final_dict_product)
        self.excel_table_create_and_insert(dict_=final_dict_product)
        time.sleep(self.parameters["wait_time"])
        if self.parameters["num_pages"] > 2:
            for i in range(2, self.parameters["num_pages"]):
                self.__driver.get(first_page_link + f"?p={i}")
                final_dict_product = self.parse_content_on_page(self.__driver.page_source)
                self.dict_to_json(dict_=final_dict_product)
                self.excel_table_create_and_insert(dict_=final_dict_product)
        self.__driver.close()   
        # except Exception as e: 
        #     print(e, flush=True) 
        #     LOGGER.critical(f"Функция run была остановлена по причине ошибки {e} {datetime.now()}")

    def dict_to_json(self, dict_: Mapping) -> None:
        """
        Преобразует данные в json файл
        """
        file_name = file_name_datetime()
        with open(f'data/output/{file_name}.json', 'w', encoding='utf-8') as file:
            file.write(json.dumps(dict_))

    def excel_table_create_and_insert(self, dict_: Mapping, page: int = 0) -> None:
        """
        Преобразует данные в excel файл
        """
        letters = list(ascii_uppercase)
        wb = Workbook()
        if page > 2:
            ws = wb.create_sheet(str(page))
        else: 
            ws = wb.active
        try:
            for index in range(0, len(list(dict_.values()))+1):
                index_row = 0
                for key, value in dict_.items():
                    letter = letters[index_row]
                    index_column = 2
                    print(key)
                    ws[f"{letter}1"] = key
                    for sub_value in value:
                        print(sub_value)
                        ws[f"{letter}{index_column}"] = sub_value
                        print(f"{letter}{index_column}")
                        index_column += 1 
                    index_row += 1
            file_name = file_name_datetime()
            wb.save(f"data/output/{file_name}.xlsx")
        except Exception as e:
            pass
        finally:
            file_name = file_name_datetime()
            wb.save(f"data/output/{file_name}.xlsx")

    def format_elements_name(self, list_: Iterable) -> Iterable:
        """
        Форматирует строковые значения элементов, что бы лишнее не попадало в общий список
        """
        for i in list_:
            if i in serialize_data_areas() or i in get_categories():
                list_ = list_.remove(i)
        return list_


    def parse_content_on_page(self, page_source: str) -> Iterable:
        """
        Парсит страницу с помощью beatifulsoup
        """
        soup = BeautifulSoup(page_source, 'html.parser')
        final_dict = {}
        if self.parameters['url']:
            elements_url = [str(i.get('href')) for i in soup.find_all(attrs={'itemprop': 'url'})]
            elements_url = self.delete_duplicate_links(elements_url)
            final_dict["elements_url"] = elements_url
        if self.parameters['title']:
            elements_name = [str(i.text) for i in soup.find_all(attrs={'itemprop': 'name'})]
            elements_name = self.format_elements_name(elements_name)
            final_dict["elements_title"] = elements_name
        if self.parameters['description']:
            elements_description = [str(i.get('content')) for i in soup.find_all(attrs={'itemprop': 'description'})]
            final_dict["elements_description"] = elements_description
        if self.parameters['price']:
            elements_price = [str(i.get('content')) for i in soup.find_all(attrs={'itemprop': "price"})]
            elements_price_currency = [i.get('content') for i in soup.find_all(attrs={'itemprop': "priceCurrency"})]
            final_dict["price"] = [x + y for (x, y) in zip(elements_price, elements_price_currency)]
        if self.parameters['paid_goods_used']:
            paid_goods = [not i.find('i') is None for i in soup.find_all(attrs={'data-marker': 'item'})]
            paid_goods = list(map(lambda x: str(x), paid_goods))
            final_dict["paid_goods"] = paid_goods 
        if self.parameters['datetime_']:
            dates = [str(i.get_text()) for i in soup.find_all(attrs={'data-marker': "item-date"})]
            final_dict['datetime_'] = dates
        if self.parameters['views_num']:
            final_dict["views_num"] = self.get_views_from_page(page_source=page_source)
        return final_dict

    def format_text(self, string_: str) -> str:
        """
        Форматирует текст
        """
        for i in string_:
            if i in list(string.ascii_letters + """1234567890<>="'()/-;._,"""):
                string_ = string_.replace(i, '')
        return string_

    def format_text_int(self, string_: str) -> str:
        """
        Форматирует текст в целочисленное значение
        """
        for i in string_:
            if str(i) not in list("""1234567890"""):
                string_ = string_.replace(i, '')
        return string_


    def get_views_and_date_from_page(self, page_source: str) -> Mapping:
        """
        Парсит и просмотры, и дату, что бы не было лишних запросов для одной страницы
        """
        soup = BeautifulSoup(page_source, 'html.parser')
        elements_url = [str(i.get('href')) for i in soup.find_all(attrs={'itemprop': 'url'})]
        final_list_views = []
        final_list_datetime_ = []
        elements_url = self.delete_duplicate_links(elements_url)
        for i in elements_url:
            time.sleep(self.parameters['wait_time'])
            self.__driver.get(self.__LINK + i)
            source_page = self.__driver.page_source
            soup = BeautifulSoup(source_page, 'html.parser')
            date_time = soup.find(attrs={"data-marker":'item-view/item-date'})
            total = self.format_text_int(soup.find(attrs={"data-marker": "item-view/total-views"}).text)
            today = self.format_text_int(soup.find(attrs={"data-marker": "item-view/today-views"}).text)
            final_list_views.append(f"Общее количество просмотров - {total}, Количество просмотров сегодня - {today}")
            final_list_datetime_.append(date_time)
        return {
            "final_list_views": final_list_views,
            "final_list_datetime_": final_list_datetime_
        }


    def get_views_from_page(self, page_source: str) -> Iterable:
        """
        Парсит количество просмотров на странице
        """
        soup = BeautifulSoup(page_source, 'html.parser')
        elements_url = [str(i.get('href')) for i in soup.find_all(attrs={'data-marker': 'item-title'})]
        elements_url = self.delete_duplicate_links(elements_url)
        final_list = []
        for i in elements_url:
            time.sleep(self.parameters['wait_time'])
            self.__driver.get(self.__LINK + i)
            source_page = self.__driver.page_source
            soup = BeautifulSoup(source_page, 'html.parser')
            total = self.format_text(soup.find(attrs={"data-marker": "item-view/total-views"}).text)
            today = self.format_text(soup.find(attrs={"data-marker": "item-view/today-views"}).text())
            final_list.append(f"Общее количество просмотров - {total}, Количество просмотров сегодня - {today}")
        return final_list


    def delete_duplicate_links(self, list_: Iterable) -> Iterable:
        """
        Удаляем дубликаты из списка ссылок
        """
        for inner in list_:
            if list_.count(inner) == 2:
                list_.remove(inner)
        return list_


    def get_date_time_from_page(self, page_source: str) -> Iterable:
        """
        Парсит дату на странице
        """
        final_list = []
        soup = BeautifulSoup(page_source, 'html.parser')
        elements_url = [i.get('href') for i in soup.find_all(attrs={'data-marker': 'item-title'})]
        elements_url = self.delete_duplicate_links(elements_url)
        for i in elements_url:
            self.__driver.get(self.__LINK + i)
            source_page = self.__driver.page_source
            soup = BeautifulSoup(source_page, 'html.parser')
            date_time = self.format_text(soup.find(attrs={"data-marker":'item-view/item-date'}).text)
            final_list.append(date_time)
        return final_list


    def kill(self):
        """
        Закрывает Селениум
        """
        LOGGER.info(f"Функция kill запущена {datetime.now()}") 
        try:
            self.__driver.close() 
        except Exception as e:
            print(e, flush=True)
            LOGGER.critical(f"Функция kill была остановлена по причине ошибки {e} {datetime.now()}")


    def __make_clicks_to_get_right_city(self):
        """
        Делает клики для того, что бы получить правильный город
        """
        LOGGER.info(f"Функция __make_clicks_to_get_right_city запущена {datetime.now()}") 
        try:
            self.__driver.find_element(By.XPATH, self.city__xpath_of_elements()["city_choose"]).click()
            time.sleep(self.parameters["wait_time"])
            self.__driver.find_element(By.XPATH, self.city__xpath_of_elements()["send_form_city_choose"]).send_keys(self.city)
            time.sleep(self.parameters["wait_time"])
            self.__driver.find_element(By.XPATH, self.city__xpath_of_elements()["first_element_city_choose"]).click()
            time.sleep(self.parameters["wait_time"])
            self.__driver.find_element(By.XPATH, self.city__xpath_of_elements()["at_first_check_this_town's_posts"]).click()
            time.sleep(self.parameters["wait_time"])
            self.__driver.find_element(By.XPATH, self.city__xpath_of_elements()["confirm_city_choose"]).click()
        except Exception as e:
            print(e, flush=True)
            LOGGER.critical(f"Функция run была остановлена по причине ошибки {e} {datetime.now()}")


    def __only_name_choose(self):
        LOGGER.info(f"Функция __only_name_choose запущена {datetime.now()}")
        try: 
            if self.parameters['name_only']:
                only_name_dict = {
                    "only_name_check_box": """//*[@id="app"]/div/div[2]/div/div[2]/div/div[3]/button""",
                    "submit_button": """//*[@id="app"]/div/div[2]/div/div[2]/div/div[3]/button""",
                }
                self.__driver.find_element(By.XPATH, only_name_dict["only_name_check_box"]).click()
                self.__driver.find_element(By.XPATH, only_name_dict["submit_button"]).click()
            else:
                pass
        except Exception as e:
            print(e, flush=True)
            LOGGER.critical(f"Функция __only_name_choose была остановлена по причине ошибки {e} {datetime.now()}")


    def __make_clicks_to_get_category(self):
        """
        Делает клики для получения категории
        """
        LOGGER.info(f"Функция __make_clicks_to_get_category запущена {datetime.now()}") 
        try:
            Select(self.__driver.find_element(By.XPATH, self.category__xpath_of_elemets()["category_choose"])).select_by_visible_text(self.parameters["category"])
        except Exception as e:
            print(e)
            LOGGER.critical(f"Функция __make_clicks_to_category была остановлена по причине ошибки {e} {datetime.now()}")


    def city__xpath_of_elements(self) -> Mapping:
        """
        Возвращает словарь на ссылки городов в формате xpath
        """
        LOGGER.info(f"Функция city__xpath_of_elements запущена {datetime.now()}") 
        dict_of_city_xpathes = {
            "city_choose": """//*[@id="app"]/div/div[2]/div/div[2]/div/div[6]/div/div/span/div/div[1]/div[2]/div""",
            "first_element_city_choose": """//*[@id="app"]/div/div[2]/div/div[2]/div/div[6]/div/div/span/div/div[1]/div[2]/div/ul/li[1]""",
            "send_form_city_choose": """//*[@id="app"]/div/div[2]/div/div[2]/div/div[6]/div/div/span/div/div[1]/div[2]/div/input""",
            "at_first_check_this_town's_posts": """//*[@id="app"]/div/div[2]/div/div[2]/div/div[6]/div/div/span/div/div[3]/div/div[1]/label/span""",
            "confirm_city_choose": """//*[@id="app"]/div/div[2]/div/div[2]/div/div[6]/div/div/span/div/div[3]/div/div[2]/div/button"""
        }
        return dict_of_city_xpathes 


    def category__xpath_of_elemets(self) -> Mapping:
        """
        Возвращает словарь на ссылки категорий в формате xpath
        """
        LOGGER.info(f"Функция category__xpath_of_elements запущена {datetime.now()}") 
        try:
            dict_of_category_xpathes = {
                "category_choose": """//*[@id="category"]"""
            }
            return dict_of_category_xpathes
        except Exception as e:
            print(e)
            LOGGER.critical(f"Функция category__xpath_of_elements была остановлена по причине ошибки {e} {datetime.now()}")
            

